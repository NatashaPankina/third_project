from flask import Flask, render_template, redirect, request, abort, url_for, send_file
from data import db_session
from data.schools import School
from data.users import User
from data.students import Student
from data.olympiads import Olympiad
from flask_login import LoginManager, login_user, login_required, current_user, logout_user
import openpyxl as xl
import os


app = Flask(__name__)
app.config['SECRET_KEY'] = 'svetik'
DOWNLOAD_FOLDER = os.path.dirname(os.path.abspath(__file__)) + '/downloads/'
app.config['DOWNLOAD_FOLDER'] = DOWNLOAD_FOLDER
login_manager = LoginManager()
login_manager.init_app(app)  

db_session.global_init("db/all.db")
db_sess = db_session.create_session()
school = School()
school.name = 'не выбрано'
olympiad = Olympiad()
olympiad.name = 'не выбрано'
try:
    db_sess.add(olympiad)
    db_sess.commit()
    db_sess.add(school)
    db_sess.commit()
except Exception:
    pass
res_f = []


def from_db_to_excel(values):
    wb = xl.Workbook()
    ws = wb.active
    ws.title = "Олимпиадники"
    for i in values:
        ws.append(i)
    wb.save('olimp_results.xlsx')
    return 'olimp_results.xlsx'

    
@app.route('/', methods=['GET', 'POST'])
def start_page():
    return render_template('first.html', title="Олимпиады и олимпиадники")


@app.route('/uploads/<filename>', methods=['GET', 'POST'])
def download(filename):
    filename = from_db_to_excel(res_f)
    #uploads = os.path.join(app.root_path, app.config['DOWNLOAD_FOLDER'])
    return send_file(filename, as_attachment=True)


@app.route('/konvert_output', methods=['GET', 'POST'])
@login_required
def konvert_output():
    if request.method == 'GET':
        return render_template('konvert.html', name=f'{current_user.surname} {current_user.name} {current_user.patronymic}', text='Колонки таблицы должны полностью совпадать с таблицей в разделе "Ученики" за исключением первой колонки, её просто не должно быть', title="Конвертация из Excel")
    if request.method == 'POST':
        try:
            f = request.files['file'].filename
            k = xl.load_workbook(f)
            for i in k.sheetnames:
                data = k[i]
                for row in data.values:
                    if 'Фамилия' not in row:
                        db_sess = db_session.create_session()
                        student = Student()
                        student.surname = row[0]
                        student.name = row[1]
                        student.patronymic = row[2]
                        student.class_writing = row[3]
                        student.class_take = row[4]
                        student.school_id = db_sess.query(School.id).filter(row[5] == School.name).first()[0]
                        student.status = row[6].lower()
                        student.olymp_id = db_sess.query(Olympiad.id).filter(row[7] == Olympiad.name).first()[0]
                        f, s, t = row[8].split()
                        student.user_id = db_sess.query(User.id).filter(f == User.surname, s == User.name, t == User.patronymic).first()[0]
                        student.year = row[9]
                        db_sess.add(student)
                        db_sess.commit()
        except Exception as e:
            print(e)
            return render_template('err.html', message="Проверьте корректность файла", title="Ошибка")
        return render_template('konvert.html', name=f'{current_user.surname} {current_user.name} {current_user.patronymic}', text='Файл успешно загружен', title="Конвертация из Excel")

    
@login_manager.user_loader
def load_user(id):
    db_sess = db_session.create_session()
    return db_sess.query(User).get(id)


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        db_sess = db_session.create_session()
        user = db_sess.query(User).filter(User.email == request.form.get('username')).first()
        if user and user.check_password(request.form.get('password')):
            login_user(user)
            return redirect("/students")
        return render_template('login.html', message="Неправильный логин или пароль")
    return render_template('login.html', title='Вход')


@app.route('/students', methods=['GET', 'POST'])
@login_required
def success():
    global res_f

    db_sess = db_session.create_session()
    if current_user.admin:
        res1 = db_sess.query(School.name).distinct().all()
        res3 = db_sess.query(User.surname, User.name, User.patronymic).distinct().all()
    else:
        res1 = [(current_user.school.name,)]
        res3 = [(f'{current_user.surname} {current_user.name} {current_user.patronymic}',)]
    res2 = db_sess.query(Olympiad.name).distinct().all()
    res4 = db_sess.query(Student.year).distinct().all()
    res5 = db_sess.query(Student.class_writing).distinct().all()
    for i in range(len(res3)):
        res3[i] = ' '.join(list(res3[i]))
    for i in range(len(res1)):
        res1[i] = res1[i][0]
    for i in range(len(res2)):
        res2[i] = res2[i][0]
    for i in range(len(res4)):
        res4[i] = res4[i][0]
    for i in range(len(res5)):
        res5[i] = res5[i][0]
    if True:
        if current_user.admin:
            db_sess = db_session.create_session()
            res_f = db_sess.query(Student.id, Student.surname, Student.name, Student.patronymic, Student.class_writing, Student.class_take, Student.school_id, Student.status, Student.olymp_id, Student.user_id, Student.year).all()
            for i in range(len(res_f)):
                res_f[i] = list(res_f[i])
                res_f[i][6] = db_sess.query(School.name).filter(res_f[i][6] == School.id).first()[0]
                res_f[i][8] = db_sess.query(Olympiad.name).filter(res_f[i][8] == Olympiad.id).first()[0]
                res_f[i][9] = ' '.join(list(db_sess.query(User.surname, User.name, User.patronymic).filter(res_f[i][9] == User.id).first()))
        else:
            b_sess = db_session.create_session()
            res_f = db_sess.query(Student.id, Student.surname, Student.name, Student.patronymic, Student.class_writing, Student.class_take, Student.school_id, Student.status, Student.olymp_id, Student.user_id, Student.year).filter(Student.user_id == current_user.id).all()
            for i in range(len(res_f)):
                res_f[i] = list(res_f[i])
                res_f[i][6] = db_sess.query(School.name).filter(res_f[i][6] == School.id).first()[0]
                res_f[i][8] = db_sess.query(Olympiad.name).filter(res_f[i][8] == Olympiad.id).first()[0]
                res_f[i][9] = ' '.join(list(db_sess.query(User.surname, User.name, User.patronymic).filter(res_f[i][9] == User.id).first()))
        id2 = request.form.get('new_id')
        if request.method == 'POST':
            school = db_sess.query(School.id).filter(School.name.in_(request.form.getlist('type1'))).all()
            for i in range(len(school)):
                school[i] = school[i][0]
            classs = request.form.getlist('type2')
            for i in range(len(classs)):
                classs[i] = classs[i][0]
            teacher = request.form.getlist('type3')
            for i in range(len(teacher)):
                f, s, t = teacher[i].split()
                teacher[i] = db_sess.query(User.id).filter(f == User.surname, s == User.name, t == User.patronymic).first()[0]
            olympiad = db_sess.query(Olympiad.id).filter(Olympiad.name.in_(request.form.getlist('type4'))).all()
            for i in range(len(olympiad)):
                olympiad[i] = olympiad[i][0]
            year = request.form.getlist('type5')
            for i in range(len(year)):
                year[i] = year[i]
            if len(school) == 0:
                school = db_sess.query(School.id).distinct().all()
                for i in range(len(school)):
                    school[i] = school[i][0]                
            if len(classs) == 0:
                classs = res5
            if len(teacher) == 0:
                teacher = db_sess.query(User.id).distinct().all()
                for i in range(len(teacher)):
                    teacher[i] = teacher[i][0]
            if len(olympiad) == 0:
                olympiad = db_sess.query(Olympiad.id).distinct().all()
                for i in range(len(olympiad)):
                    olympiad[i] = olympiad[i][0]
            if len(year) == 0:
                year = res4

                

            res_f = db_sess.query(Student.id, Student.surname, Student.name, Student.patronymic, Student.class_writing, Student.class_take, Student.school_id, Student.status, Student.olymp_id, Student.user_id, Student.year).filter(Student.class_writing.in_(classs), Student.school_id.in_(school), Student.olymp_id.in_(olympiad), Student.year.in_(year), Student.user_id.in_(teacher)).all()
            if not(current_user.admin):
                res_f = []
                for i in res_f:
                    i = list(i)
                    if current_user.id == i[9]:
                        res_f.append(tuple(i))

            for i in range(len(res_f)):
                res_f[i] = list(res_f[i])
                res_f[i][6] = db_sess.query(School.name).filter(res_f[i][6] == School.id).first()[0]
                res_f[i][8] = db_sess.query(Olympiad.name).filter(res_f[i][8] == Olympiad.id).first()[0]
                res_f[i][9] = ' '.join(list(db_sess.query(User.surname, User.name, User.patronymic).filter(res_f[i][9] == User.id).first()))

    return render_template("main.html", title="Ученики", name=f'{current_user.surname} {current_user.name} {current_user.patronymic}', bd=res_f, teachers=res3, schools=res1, classes=res5, years=res4, olympiads=res2, id=id2)


@app.route('/register', methods=['GET', 'POST'])
def reqister():
    db_sess = db_session.create_session()
    res = db_sess.query(School.name).all()
    if request.method == 'POST':
        if len(request.form.get('surname')) == 0 or len(request.form.get('name')) == 0 or len(request.form.get('patronymic')) == 0:
            return render_template('register.html', title='Регистрация',
                                   message="Не все обязательные поля заполнены", bd=res)
        if request.form.get('password') != request.form.get('password_again'):
            return render_template('register.html', title='Регистрация',
                                   message="Пароли не совпадают", bd=res)
        db_sess = db_session.create_session()
        if db_sess.query(User).filter(User.email == request.form.get('username')).first():
            return render_template('register.html', title='Регистрация',
                                  message="Такой пользователь уже есть", bd=res)
        sch = db_sess.query(School.id).filter(request.form.get('schools') == School.name).first()
        
        if request.form.get('admin') == '032803':
            user = User(
                surname=request.form.get('surname'),
                name=request.form.get('name'),
                patronymic=request.form.get('patronymic'),
                email=request.form.get('username'),
                school_id=sch[0],
                admin=True)
        else:
            user = User(
                surname=request.form.get('surname'),
                name=request.form.get('name'),
                patronymic=request.form.get('patronymic'),
                email=request.form.get('username'),
                school_id=sch[0],
                admin=False)
        user.set_password(request.form.get('password'))
        db_sess.add(user)
        db_sess.commit()
        return redirect('/login')
    return render_template('register.html', title='Регистрация', bd=res)


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect("/")


@app.route('/students/add', methods=['GET', 'POST'])
@login_required
def add():
    surname = ''
    name = ''
    patronymic = ''
    class_writing = ''
    class_take = ''
    year = ''
    db_sess = db_session.create_session()
    if current_user.admin:
        res1 = db_sess.query(School.name).all()
        res3 = db_sess.query(User.surname, User.name, User.patronymic).all()
    else:
        res1 = [(current_user.school.name,)]
        res3 = [(f'{current_user.surname} {current_user.name} {current_user.patronymic}',)]
    res2 = db_sess.query(Olympiad.name).all()
    for i in range(len(res3)):
        res3[i] = (' '.join(list(res3[i])),)
    
    if request.method == 'POST':
        db_sess = db_session.create_session()
        student = Student()
        student.surname = request.form.get('surname')
        student.name = request.form.get('name')
        student.patronymic = request.form.get('patronymic')
        student.class_writing = request.form.get('class_writing')
        student.class_take = request.form.get('class_take')
        student.school_id = db_sess.query(School.id).filter(request.form.get('schools') == School.name).first()[0]
        student.olymp_id = db_sess.query(Olympiad.id).filter(request.form.get('olympiads') == Olympiad.name).first()[0]
        student.status = request.form.get('status')
        f, s, t = request.form.get('teachers').split()
        student.user_id = db_sess.query(User.id).filter(f == User.surname, s == User.name, t == User.patronymic).first()[0]
        student.year = request.form.get('year')
        db_sess.merge(current_user)
        db_sess.add(student)
        db_sess.commit()
        return redirect("/students")
    return render_template('Addform.html', surname=surname, name=name, patronymic=patronymic, class_writing=class_writing,
                           class_take=class_take, year=year, title='Добавление', sch=res1, stat=['участник', 'призёр', 'победитель'], olim=res2, teach=res3)


@app.route('/students/<int:id>', methods=['GET', 'POST'])
@login_required
def edit(id):
    if request.method == "GET":
        db_sess = db_session.create_session()        
        if current_user.admin:
            res = db_sess.query(Student).filter(Student.id == id).first()
        else:
            res = db_sess.query(Student).filter(Student.id == id,
                                           current_user.id == Student.user_id).first()
        if res:
            surname = res.surname
            name = res.name
            patronymic = res.patronymic
            class_writing = res.class_writing
            class_take = res.class_take
            year = res.year
            sch = [(res.school.name,)] + db_sess.query(School.name).filter(School.name != res.school.name).all()
            olim = [(res.olympiad.name,)] + db_sess.query(Olympiad.name).filter(Olympiad.name != res.olympiad.name).all()
            r = ['участник', 'призёр', 'победитель']
            r.remove(res.status)
            st = [res.status] + r
            if current_user.admin:
                teach = [(res.user.surname, res.user.name, res.user.patronymic)] + db_sess.query(User.surname, User.name, User.patronymic).filter(User.surname != res.user.surname, User.name != res.user.name, User.patronymic != res.user.patronymic).all()
                for i in range(len(teach)):
                    teach[i] = (' '.join(list(teach[i])),)
            else:
                teach = [(current_user.surname, current_user.name, current_user.patronymic)]
        else:
            return render_template('err.html', message="Проверьте корректность запроса", title="Ошибка")

    if request.method == "POST":
        db_sess = db_session.create_session()
        res = db_sess.query(Student).filter(Student.id == id).first()
        if res:
            res.surname = request.form.get('surname')
            res.name = request.form.get('name')
            res.patronymic = request.form.get('patronymic')
            res.class_writing = request.form.get('class_writing')
            res.class_take = request.form.get('class_take')
            res.year = request.form.get('year')
            res.school_id = db_sess.query(School.id).filter(request.form.get('schools') == School.name).first()[0]
            res.olymp_id = db_sess.query(Olympiad.id).filter(request.form.get('olympiads') == Olympiad.name).first()[0]
            res.status = request.form.get('status')
            f, s, t = request.form.get('teachers').split()
            res.user_id = db_sess.query(User.id).filter(f == User.surname, s == User.name, t == User.patronymic).first()[0]
            res.year = request.form.get('year')
            db_sess.commit()
            return redirect("/students")
        else:
            return render_template('err.html', message="Проверьте корректность запроса", title="Ошибка")
    return render_template('Addform.html',
                           title='Редактирование', surname=surname, name=name, patronymic=patronymic, class_writing=class_writing,
                           class_take=class_take, sch=sch, stat=st, olim=olim, teach=teach, year=year)


@app.route('/students/delete/<int:id>', methods=['GET', 'POST'])
@login_required
def delete(id):
    db_sess = db_session.create_session()
    if current_user.admin:
        res = db_sess.query(Student).filter(Student.id == id).first()
    else:
        res = db_sess.query(Student).filter(Student.id == id, current_user.id == Student.user_id).first()
    if res:
        db_sess.delete(res)
        db_sess.commit()
    else:
        return render_template('err.html', message="Проверьте корректность запроса", title="Ошибка")
    return redirect("/students")


@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    if True:
        db_sess = db_session.create_session()        
        surname = current_user.surname
        name2 = current_user.name
        patronymic = current_user.patronymic
        sch = [(current_user.school.name,)] + db_sess.query(School.name).filter(School.name != current_user.school.name).all()
        username = current_user.email

    if request.method == "POST":
        db_sess = db_session.create_session()
        res = db_sess.query(User).filter(User.id == current_user.id).first()
        if res:
            res.surname = request.form.get('surname')
            res.name = request.form.get('name')
            res.patronymic = request.form.get('patronymic')
            res.school_id = db_sess.query(School.id).filter(request.form.get('schools') == School.name).first()[0]
            p = db_sess.query(User.email).filter(User.email == request.form.get('username')).first()
            if p is None or p[0] == current_user.email:
                res.email = request.form.get('username')
            else:
                return render_template('profile.html',
                                       title='Профиль', surname=surname, name2=name2, patronymic=patronymic, sch=sch, username=username, name=f'{current_user.surname} {current_user.name} {current_user.patronymic}', text='Данный логин уже используется')
            if len(request.form.get('password').strip()) != 0:
                res.set_password(request.form.get('password'))
            db_sess.commit()
            return redirect("/students")
        else:
            return render_template('err.html', message="Проверьте корректность запроса", title="Ошибка")
    return render_template('profile.html',
                           title='Профиль', surname=surname, name2=name2, patronymic=patronymic, sch=sch, username=username, name=f'{current_user.surname} {current_user.name} {current_user.patronymic}')


@app.route('/schools/add', methods=['GET', 'POST'])
@login_required
def add2():
    name = f'{current_user.surname} {current_user.name} {current_user.patronymic}'
    db_sess = db_session.create_session()
    if current_user.admin:
        if request.method == 'POST':
            db_sess = db_session.create_session()
            school = School()
            school.name = request.form.get('name')
            try:
                db_sess.merge(current_user)
                db_sess.add(school)
                db_sess.commit()
            except Exception:
                return render_template('Addschool.html', name=name, title="Добавление школы", message="Данная школа уже добавлена")
            return render_template('Addschool.html', name=name, title="Добавление школы", message="Школа успешно добавлена")
    else:
        return render_template('err.html', message="Проверьте корректность запроса", title="Ошибка")
    return render_template('Addschool.html', name=name, title="Добавление школы")


@app.route('/olympiad/add', methods=['GET', 'POST'])
@login_required
def add3():
    name = f'{current_user.surname} {current_user.name} {current_user.patronymic}'
    db_sess = db_session.create_session()
    if current_user.admin:
        if request.method == 'POST':
            db_sess = db_session.create_session()
            olympiad = Olympiad()
            olympiad.name = request.form.get('name')
            try:
                
                db_sess.merge(current_user)
                db_sess.add(olympiad)
                db_sess.commit()
            except Exception:
                return render_template('Addolimpiad.html', name=name, title="Добавление олимпиады", message="Данная олимпиада уже добавлена")
            return render_template('Addolimpiad.html', name=name, title="Добавление олимпиады", message="Олимпиада успешно добавлена")
    else:
        return render_template('err.html', message="Проверьте корректность запроса", title="Ошибка")
    
    return render_template('Addolimpiad.html', name=name, title="Добавление олимпиады")


if __name__ == '__main__':
    app.run(port=8080, host='127.0.0.1')